"""
Excel document extractor for .xlsx and .xls files
"""
import time
import io
from typing import List
from .base_extractor import BaseExtractor, ExtractionResult


class ExcelExtractor(BaseExtractor):
    """Excel spreadsheet text extraction"""
    
    @property
    def supported_mime_types(self) -> List[str]:
        return [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
            'application/excel',
            'application/x-excel'
        ]
    
    @property
    def supported_extensions(self) -> List[str]:
        return ['xlsx', 'xls']
    
    def extract_text(self, file_data: bytes, filename: str = "") -> ExtractionResult:
        """
        Extract text from Excel files
        
        Args:
            file_data: Excel file bytes
            filename: Original filename
            
        Returns:
            ExtractionResult with extracted text and metadata
        """
        start_time = time.time()
        
        # Validate file size
        self.validate_file_size(file_data, max_size_mb=50)
        
        # Determine Excel format
        is_xlsx = filename.lower().endswith('.xlsx') or b'PK' in file_data[:10]
        
        try:
            if is_xlsx:
                return self._extract_xlsx(file_data, filename, start_time)
            else:
                return self._extract_xls(file_data, filename, start_time)
                
        except Exception as e:
            raise Exception(f"Excel processing failed: {str(e)}")
    
    def _extract_xlsx(self, file_data: bytes, filename: str, start_time: float) -> ExtractionResult:
        """Extract text from .xlsx files using openpyxl"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
            
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_data), data_only=True)
            
            all_text = ""
            sheet_count = 0
            cell_count = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = f"\n=== Sheet: {sheet_name} ===\n"
                
                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row > 0 and max_col > 0:
                    # Extract data row by row
                    rows_with_data = 0
                    
                    for row_num in range(1, min(max_row + 1, 1001)):  # Limit to 1000 rows
                        row_data = []
                        has_data = False
                        
                        for col_num in range(1, min(max_col + 1, 101)):  # Limit to 100 columns
                            cell = sheet.cell(row=row_num, column=col_num)
                            cell_value = cell.value
                            
                            if cell_value is not None:
                                # Convert to string and clean
                                cell_str = str(cell_value).strip()
                                if cell_str:
                                    row_data.append(cell_str)
                                    has_data = True
                                    cell_count += 1
                                else:
                                    row_data.append("")
                            else:
                                row_data.append("")
                        
                        if has_data:
                            # Join row data with separators
                            sheet_text += " | ".join(row_data) + "\n"
                            rows_with_data += 1
                    
                    if rows_with_data > 0:
                        all_text += sheet_text
                        sheet_count += 1
                        print(f"  Sheet '{sheet_name}': {rows_with_data} rows, {cell_count} cells")
            
            workbook.close()
            
            processing_time = time.time() - start_time
            
            metadata = {
                'file_type': 'excel_xlsx',
                'filename': filename,
                'sheets_processed': sheet_count,
                'total_cells': cell_count,
                'processing_time_seconds': processing_time,
                'extractor': 'openpyxl'
            }
            
            print("Excel (.xlsx) processing completed:")
            print(f"  - Sheets: {sheet_count}")
            print(f"  - Cells: {cell_count}")
            print(f"  - Text length: {len(all_text)} characters")
            print(f"  - Processing time: {processing_time:.2f}s")
            
            return ExtractionResult(
                text=all_text,
                pages=sheet_count,
                metadata=metadata,
                processing_time=processing_time
            )
            
        except ImportError:
            raise Exception("openpyxl not installed - cannot process .xlsx files")
        except InvalidFileException:
            raise Exception("Invalid or corrupted Excel file")
        except Exception as e:
            raise Exception(f"XLSX extraction failed: {str(e)}")
    
    def _extract_xls(self, file_data: bytes, filename: str, start_time: float) -> ExtractionResult:
        """Extract text from .xls files using xlrd"""
        try:
            import xlrd
            from xlrd.biffh import XLRDError
            
            # Load workbook from bytes
            workbook = xlrd.open_workbook(file_contents=file_data)
            
            all_text = ""
            sheet_count = 0
            cell_count = 0
            
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                sheet_text = f"\n=== Sheet: {sheet_name} ===\n"
                
                rows_with_data = 0
                
                for row_num in range(min(sheet.nrows, 1000)):  # Limit to 1000 rows
                    row_data = []
                    has_data = False
                    
                    for col_num in range(min(sheet.ncols, 100)):  # Limit to 100 columns
                        try:
                            cell = sheet.cell(row_num, col_num)
                            cell_value = cell.value
                            
                            if cell_value is not None and str(cell_value).strip():
                                # Handle different cell types
                                if cell.ctype == xlrd.XL_CELL_DATE:
                                    # Convert date to string
                                    date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                                    cell_str = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                                else:
                                    cell_str = str(cell_value).strip()
                                
                                if cell_str:
                                    row_data.append(cell_str)
                                    has_data = True
                                    cell_count += 1
                                else:
                                    row_data.append("")
                            else:
                                row_data.append("")
                                
                        except Exception:
                            row_data.append("")
                    
                    if has_data:
                        sheet_text += " | ".join(row_data) + "\n"
                        rows_with_data += 1
                
                if rows_with_data > 0:
                    all_text += sheet_text
                    sheet_count += 1
                    print(f"  Sheet '{sheet_name}': {rows_with_data} rows, {cell_count} cells")
            
            processing_time = time.time() - start_time
            
            metadata = {
                'file_type': 'excel_xls',
                'filename': filename,
                'sheets_processed': sheet_count,
                'total_cells': cell_count,
                'processing_time_seconds': processing_time,
                'extractor': 'xlrd'
            }
            
            print("Excel (.xls) processing completed:")
            print(f"  - Sheets: {sheet_count}")
            print(f"  - Cells: {cell_count}")
            print(f"  - Text length: {len(all_text)} characters")
            print(f"  - Processing time: {processing_time:.2f}s")
            
            return ExtractionResult(
                text=all_text,
                pages=sheet_count,
                metadata=metadata,
                processing_time=processing_time
            )
            
        except ImportError:
            raise Exception("xlrd not installed - cannot process .xls files")
        except XLRDError as e:
            raise Exception(f"Invalid or corrupted Excel file: {str(e)}")
        except Exception as e:
            raise Exception(f"XLS extraction failed: {str(e)}")